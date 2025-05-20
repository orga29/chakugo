import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from datetime import datetime, timedelta, timezone # timedelta が使われていることを確認

# --- Excel 抽出・書式設定関数 ---
def process_workbook(df):
    # 抽出条件
    col_code = '商品コード'
    col_name = '商品名'
    col_type = '箱/こもの'
    col_qty = '集荷便から降ろす数/小分けしないと足りない数'

    # フィルタリング
    mask_qty_neg = df[col_qty] < 0
    mask_type = df[col_type].astype(str).str.contains('こもの')
    mask_no_marker = ~df[col_name].astype(str).str.endswith('◇')
    mask_exclude_higashiichi = ~df[col_name].astype(str).str.endswith('東一')
    filtered = df[mask_qty_neg & mask_type & mask_no_marker & mask_exclude_higashiichi].copy()
    filtered['必要数'] = filtered[col_qty].abs()

    # Workbook 作成
    wb = Workbook()
    ws = wb.active
    ws.title = '抽出結果'

    # ヘッダー
    ws.append([col_code, col_name, '必要数'])
    # データ行
    for row in filtered.itertuples(index=False):
        ws.append([row.商品コード, row.商品名, row.必要数])

    # 書式設定
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_gray = PatternFill(fill_type='solid', fgColor='EEEEEE')
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFF')
    tot_rows = ws.max_row
    for i, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=tot_rows, min_col=1, max_col=3), start=1):
        fill = fill_gray if i % 2 == 0 else fill_white
        ws.row_dimensions[i].height = 18.75
        for cell in row_cells:
            cell.border = border
            if i == 1:
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if cell.column_letter == 'C':
                    cell.font = Font(size=12, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = fill

    # 列幅設定
    ws.column_dimensions['A'].width = 11.25
    ws.column_dimensions['B'].width = 49.00
    ws.column_dimensions['C'].width = 10.50

    # フッター追加
    footer_row = tot_rows + 2
    footer_cell = ws.cell(row=footer_row, column=2)
    # --- 変更点 1: フッターの日付を翌日に変更 ---
    dt_jst_next_day_footer = datetime.now(timezone(timedelta(hours=9))) + timedelta(days=1)
    footer_cell.value = f"{dt_jst_next_day_footer.strftime('%m/%d')} 着後必要数"
    footer_cell.font = Font(size=12)
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')

    return wb

# --- Streamlit UI ---

st.title("着後必要数 抽出ツール")
st.write("Excelファイルをドラッグ&ドロップまたは選択すると、抽出・書式設定してダウンロードできます。")

uploaded = st.file_uploader(
    "Excelファイルをドラッグ&ドロップまたは選択 (.xlsx)",
    type="xlsx"
)
if uploaded:
    df = pd.read_excel(uploaded, header=1, skiprows=[2,3])
    wb = process_workbook(df)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # ダウンロードボタン
    # --- 変更点 2: ファイル名の日付を翌日に変更 ---
    dt_jst_next_day_filename = datetime.now(timezone(timedelta(hours=9))) + timedelta(days=1)
    st.download_button(
        label="抽出結果をダウンロード",
        data=output,
        file_name=f"{dt_jst_next_day_filename.strftime('%m%d')}着後必要数.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
